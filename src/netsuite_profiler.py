import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from netsuite_mock_data import TABLES


def build_profile(tables: dict) -> list[dict]:
    rows = []
    for table, base_count in tables.items():
        row_count = base_count + random.randint(-50, 50)
        rows.append({"table": table, "row_count": row_count})
    return rows


def compute_stats(profile: list[dict]) -> dict:
    counts = [r["row_count"] for r in profile]
    return {
        "total_tables": len(profile),
        "total_rows": sum(counts),
        "min_rows": min(counts),
        "max_rows": max(counts),
        "avg_rows": round(sum(counts) / len(counts), 1),
    }


def write_excel(profile: list[dict], stats: dict, path: str):
    wb = Workbook()

    # --- Sheet 1: Table Profile ---
    ws = wb.active
    ws.title = "Table Profile"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["Table", "Row Count"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, entry in enumerate(profile, start=2):
        ws.cell(row=row_idx, column=1, value=entry["table"])
        ws.cell(row=row_idx, column=2, value=entry["row_count"])

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14

    # --- Sheet 2: Summary Stats ---
    ws2 = wb.create_sheet("Summary")
    stat_labels = [
        ("Total Tables", stats["total_tables"]),
        ("Total Rows", stats["total_rows"]),
        ("Min Row Count", stats["min_rows"]),
        ("Max Row Count", stats["max_rows"]),
        ("Avg Row Count", stats["avg_rows"]),
    ]

    ws2.cell(row=1, column=1, value="Metric").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Value").font = Font(bold=True)

    for i, (label, value) in enumerate(stat_labels, start=2):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=value)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 14

    wb.save(path)
    print(f"Report saved to: {path}")


def main():
    profile = build_profile(TABLES)

    print(f"{'Table':<20} {'Row Count':>10}")
    print("-" * 32)
    for entry in profile:
        print(f"{entry['table']:<20} {entry['row_count']:>10,}")

    stats = compute_stats(profile)
    print()
    print(f"Total tables : {stats['total_tables']}")
    print(f"Total rows   : {stats['total_rows']:,}")
    print(f"Min rows     : {stats['min_rows']:,}")
    print(f"Max rows     : {stats['max_rows']:,}")
    print(f"Avg rows     : {stats['avg_rows']:,}")

    write_excel(profile, stats, "output/netsuite_profile.xlsx")


if __name__ == "__main__":
    main()
